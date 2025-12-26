"""
Excel dosyası işlemleri için yardımcı modül
Form verilerini Excel'den okur ve günceller
Google Sheets desteği eklendi
Bulut ortamı için optimize edildi
"""
import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Streamlit secrets desteği (bulut ortamı için)
try:
    import streamlit as st
    HAS_STREAMLIT = True
except ImportError:
    HAS_STREAMLIT = False

def get_secret(key, default=""):
    """Streamlit secrets veya environment variable'dan değer okur"""
    if HAS_STREAMLIT:
        try:
            # Streamlit secrets'tan oku
            secrets = st.secrets
            if hasattr(secrets, key):
                return getattr(secrets, key)
            # Nested secrets için (örn: secrets.google_sheets.sheet_id)
            if "." in key:
                parts = key.split(".")
                value = secrets
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    else:
                        return default
                return value
        except Exception:
            pass
    # Fallback: environment variable
    return os.environ.get(key, default)

# Google Sheets desteği (opsiyonel)
# Streamlit secrets'tan oku (nested veya flat format)
if HAS_STREAMLIT:
    try:
        secrets = st.secrets
        # Nested format kontrolü
        if hasattr(secrets, "google_sheets"):
            gs_config = secrets.google_sheets
            USE_GOOGLE_SHEETS = str(gs_config.get("enabled", "false")).lower() == "true"
            GOOGLE_SHEET_ID = str(gs_config.get("sheet_id", ""))
            GOOGLE_CREDENTIALS_JSON = str(gs_config.get("credentials_json", ""))
        else:
            # Flat format
            USE_GOOGLE_SHEETS = get_secret("USE_GOOGLE_SHEETS", "false").lower() == "true"
            GOOGLE_SHEET_ID = get_secret("GOOGLE_SHEET_ID", "")
            GOOGLE_CREDENTIALS_JSON = get_secret("GOOGLE_CREDENTIALS_JSON", "")
    except Exception:
        # Fallback to environment variables
        USE_GOOGLE_SHEETS = get_secret("USE_GOOGLE_SHEETS", "false").lower() == "true"
        GOOGLE_SHEET_ID = get_secret("GOOGLE_SHEET_ID", "")
        GOOGLE_CREDENTIALS_JSON = get_secret("GOOGLE_CREDENTIALS_JSON", "")
else:
    # Environment variables only
    USE_GOOGLE_SHEETS = get_secret("USE_GOOGLE_SHEETS", "false").lower() == "true"
    GOOGLE_SHEET_ID = get_secret("GOOGLE_SHEET_ID", "")
    GOOGLE_CREDENTIALS_JSON = get_secret("GOOGLE_CREDENTIALS_JSON", "")

# Google Apps Script URL (eski yöntem)
GOOGLE_APPS_SCRIPT_URL = get_secret("GOOGLE_APPS_SCRIPT_URL", "https://script.google.com/macros/s/AKfycbwtLKzCB366hwi1S4cHAUGIWP9dDA6isSDLbKvyOIw9P9WNgbLF6t6dlY7RYWlvQM96/exec")
USE_GOOGLE_APPS_SCRIPT = get_secret("USE_GOOGLE_APPS_SCRIPT", "true").lower() == "true"

if USE_GOOGLE_SHEETS:
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        USE_GOOGLE_SHEETS = False

def get_google_sheets_client():
    """Google Sheets client'ı oluşturur ve döndürür"""
    if not USE_GOOGLE_SHEETS:
        return None
    
    try:
        if not GOOGLE_CREDENTIALS_JSON or not GOOGLE_SHEET_ID:
            return None
        
        # JSON string'ini parse et
        import json as json_module
        if isinstance(GOOGLE_CREDENTIALS_JSON, str):
            creds_dict = json_module.loads(GOOGLE_CREDENTIALS_JSON)
        else:
            creds_dict = GOOGLE_CREDENTIALS_JSON
        
        # Credentials oluştur
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        )
        
        # Client oluştur
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        _log("ERROR", "excel_handler.py:get_google_sheets_client", "Failed to create Google Sheets client", {"error": str(e)})
        return None

# Logging - bulut ortamında devre dışı (opsiyonel olarak Streamlit logging kullanılabilir)
def _log(hypothesis_id, location, message, data):
    # Bulut ortamında logging devre dışı
    # Gerekirse Streamlit'in kendi logging sistemini kullanabilirsiniz
    pass

# Excel dosyası yolu - önce mevcut dizinde ara, yoksa geçici dizin kullan
# Google Sheets kullanılıyorsa Excel dosyası kullanılmayacak
import tempfile

# Önce mevcut dizinde ara (GitHub'dan gelen dosya için)
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_LOCAL = os.path.join(CURRENT_DIR, "form_data.xlsx")

# Geçici dizin (fallback)
TEMP_DIR = tempfile.gettempdir()
EXCEL_FILE_TEMP = os.path.join(TEMP_DIR, "form_data.xlsx")

# Önce local'de ara, yoksa temp kullan
if os.path.exists(EXCEL_FILE_LOCAL):
    EXCEL_FILE = EXCEL_FILE_LOCAL
else:
    EXCEL_FILE = EXCEL_FILE_TEMP

def create_default_excel():
    """Default değerlerle Excel dosyası oluşturur"""
    wb = Workbook()
    
    # Varsayılan sheet'i sil
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Vehicles sheet
    ws_vehicles = wb.create_sheet("Vehicles")
    ws_vehicles.append(["Vehicle"])
    ws_vehicles.append(["SPRINTER: BT-48331"])
    ws_vehicles.append(["RAM-Promaster 2500 (2021)"])
    ws_vehicles.append(["MERCEDES-2500 Cargo Van (2013)"])
    
    # FuelLevels sheet
    ws_fuel = wb.create_sheet("FuelLevels")
    ws_fuel.append(["Level"])
    ws_fuel.append(["Full"])
    ws_fuel.append(["3/4"])
    ws_fuel.append(["Half"])
    ws_fuel.append(["1/4"])
    ws_fuel.append(["Other"])
    
    # ExteriorChecks sheet
    ws_exterior = wb.create_sheet("ExteriorChecks")
    ws_exterior.append(["Field"])
    exterior_fields = ["headlights", "break_lights", "indicators", "mirrors", "windows", 
                      "windshield", "wiper_fluid", "wipers", "tires", "body_paint"]
    for field in exterior_fields:
        ws_exterior.append([field])
    
    # EngineChecks sheet
    ws_engine = wb.create_sheet("EngineChecks")
    ws_engine.append(["Field"])
    engine_fields = ["engine_noise", "coolant_level", "brake_fluid", "battery_condition", "belts_hoses"]
    for field in engine_fields:
        ws_engine.append([field])
    
    # SafetyEquipment sheet
    ws_safety = wb.create_sheet("SafetyEquipment")
    ws_safety.append(["Field"])
    safety_fields = ["seatbelts", "fire_extinguisher", "first_aid_kit", "horn", "jumper_cables", "snow_brush"]
    for field in safety_fields:
        ws_safety.append([field])
    
    # InteriorChecks sheet
    ws_interior = wb.create_sheet("InteriorChecks")
    ws_interior.append(["Field"])
    interior_fields = ["cleanliness", "dashboard_lights", "hvac", "seats"]
    for field in interior_fields:
        ws_interior.append([field])
    
    # Items sheet
    ws_items = wb.create_sheet("Items")
    ws_items.append(["Item"])
    ws_items.append(["Fuel Card"])
    ws_items.append(["Measuring Tape"])
    ws_items.append(["Safety Vest"])
    
    # Users sheet
    ws_users = wb.create_sheet("Users")
    ws_users.append(["Username", "Password", "Full Name", "Email", "Admin"])
    ws_users.append(["innovodriver", "123456", "Mehmet Berk", "mehmet.berk@example.com", "No"])
    ws_users.append(["admin", "admin123", "Admin User", "admin@example.com", "Yes"])
    
    wb.save(EXCEL_FILE)
    return wb

def get_excel_file():
    """Excel dosyasını açar, yoksa oluşturur"""
    # Google Sheets kullanılıyorsa Excel dosyasına gerek yok
    if USE_GOOGLE_SHEETS:
        # Yine de fallback için varsayılan Excel oluştur
        if not os.path.exists(EXCEL_FILE):
            return create_default_excel()
    
    if not os.path.exists(EXCEL_FILE):
        return create_default_excel()
    
    try:
        return load_workbook(EXCEL_FILE)
    except Exception as e:
        # Dosya bozuksa yeniden oluştur
        _log("E", "excel_handler.py:get_excel_file", "Excel file corrupted, recreating", {"error": str(e)})
        try:
            os.remove(EXCEL_FILE)
        except:
            pass
        return create_default_excel()

def load_vehicles():
    """Vehicles sheet'inden araç listesini okur"""
    # Google Sheets'ten oku
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Vehicles")
                all_values = sheet.get_all_values()
                vehicles = []
                for row in all_values[1:]:  # İlk satır başlık
                    if row and row[0]:
                        vehicles.append(row[0])
                return vehicles
            except Exception as e:
                _log("E", "excel_handler.py:load_vehicles:google_sheets", "Google Sheets load failed, falling back to Excel", {"error": str(e)})
    
    # Excel'den oku (fallback)
    wb = get_excel_file()
    ws = wb["Vehicles"]
    vehicles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            vehicles.append(row[0])
    return vehicles

def load_fuel_levels():
    """FuelLevels sheet'inden yakıt seviyelerini okur"""
    # Google Sheets'ten oku
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("FuelLevels")
                all_values = sheet.get_all_values()
                levels = []
                for row in all_values[1:]:  # İlk satır başlık
                    if row and row[0]:
                        levels.append(row[0])
                return levels
            except Exception as e:
                _log("E", "excel_handler.py:load_fuel_levels:google_sheets", "Google Sheets load failed, falling back to Excel", {"error": str(e)})
    
    # Excel'den oku (fallback)
    wb = get_excel_file()
    ws = wb["FuelLevels"]
    levels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            levels.append(row[0])
    return levels

def load_check_fields(category):
    """İlgili sheet'ten kontrolleri okur
    category: 'ExteriorChecks', 'EngineChecks', 'SafetyEquipment', 'InteriorChecks'
    """
    # Google Sheets'ten oku
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet(category)
                all_values = sheet.get_all_values()
                fields = []
                for row in all_values[1:]:  # İlk satır başlık
                    if row and row[0]:
                        fields.append(row[0])
                return fields
            except Exception as e:
                _log("E", f"excel_handler.py:load_check_fields:google_sheets:{category}", "Google Sheets load failed, falling back to Excel", {"error": str(e)})
    
    # Excel'den oku (fallback)
    wb = get_excel_file()
    if category not in wb.sheetnames:
        return []
    ws = wb[category]
    fields = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            fields.append(row[0])
    return fields

def load_items():
    """Items sheet'inden eşya listesini okur"""
    # Google Sheets'ten oku
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Items")
                all_values = sheet.get_all_values()
                items = []
                for row in all_values[1:]:  # İlk satır başlık
                    if row and row[0]:
                        items.append(row[0])
                return items
            except Exception as e:
                _log("E", "excel_handler.py:load_items:google_sheets", "Google Sheets load failed, falling back to Excel", {"error": str(e)})
    
    # Excel'den oku (fallback)
    wb = get_excel_file()
    ws = wb["Items"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            items.append(row[0])
    return items

def load_users():
    """Users sheet'inden kullanıcıları okur"""
    _log("A", "excel_handler.py:load_users:entry", "load_users called", {})
    
    # Google Sheets'ten oku
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Users")
                all_values = sheet.get_all_values()
                
                if not all_values or len(all_values) < 2:
                    return {}
                
                headers = all_values[0]
                _log("A", "excel_handler.py:load_users:headers", "Users sheet headers", {"headers": headers})
                
                users = {}
                for row in all_values[1:]:  # İlk satır başlık
                    if row and len(row) >= 3 and row[0] and row[1] and row[2]:
                        users[row[0]] = {
                            "password": row[1],
                            "full_name": row[2],
                            "email": row[3] if len(row) > 3 and row[3] else ""
                        }
                        _log("A", "excel_handler.py:load_users:user_added", "User added to dict", {"username": row[0]})
                
                _log("A", "excel_handler.py:load_users:exit", "load_users returning", {"user_count": len(users), "usernames": list(users.keys())})
                return users
            except Exception as e:
                _log("E", "excel_handler.py:load_users:google_sheets", "Google Sheets load failed, falling back to Excel", {"error": str(e)})
    
    # Excel'den oku (fallback)
    wb = get_excel_file()
    ws = wb["Users"]
    
    headers = [cell.value for cell in ws[1]]
    _log("A", "excel_handler.py:load_users:headers", "Users sheet headers", {"headers": headers})
    
    users = {}
    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        _log("A", "excel_handler.py:load_users:row", f"Processing row {row_num}", {"row": list(row), "row_length": len(row) if row else 0})
        if row[0] and row[1] and row[2]:
            users[row[0]] = {
                "password": row[1],
                "full_name": row[2],
                "email": row[3] if len(row) > 3 and row[3] else ""
            }
            _log("A", "excel_handler.py:load_users:user_added", "User added to dict", {"username": row[0]})
    
    _log("A", "excel_handler.py:load_users:exit", "load_users returning", {"user_count": len(users), "usernames": list(users.keys())})
    return users

def add_vehicle(vehicle_name):
    """Yeni araç ekler"""
    wb = get_excel_file()
    ws = wb["Vehicles"]
    # Aynı isimde araç var mı kontrol et
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == vehicle_name:
            return False
    ws.append([vehicle_name])
    wb.save(EXCEL_FILE)
    return True

def delete_vehicle(vehicle_name):
    """Aracı siler"""
    wb = get_excel_file()
    ws = wb["Vehicles"]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == vehicle_name:
            ws.delete_rows(row_idx)
            wb.save(EXCEL_FILE)
            return True
    return False

def update_vehicle(old_name, new_name):
    """Araç adını günceller"""
    wb = get_excel_file()
    ws = wb["Vehicles"]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == old_name:
            ws.cell(row=row_idx, column=1, value=new_name)
            wb.save(EXCEL_FILE)
            return True
    return False

def add_fuel_level(level):
    """Yeni yakıt seviyesi ekler"""
    wb = get_excel_file()
    ws = wb["FuelLevels"]
    # Aynı seviye var mı kontrol et
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == level:
            return False
    ws.append([level])
    wb.save(EXCEL_FILE)
    return True

def delete_fuel_level(level):
    """Yakıt seviyesini siler"""
    wb = get_excel_file()
    ws = wb["FuelLevels"]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == level:
            ws.delete_rows(row_idx)
            wb.save(EXCEL_FILE)
            return True
    return False

def update_fuel_level(old_level, new_level):
    """Yakıt seviyesini günceller"""
    wb = get_excel_file()
    ws = wb["FuelLevels"]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == old_level:
            ws.cell(row=row_idx, column=1, value=new_level)
            wb.save(EXCEL_FILE)
            return True
    return False

def add_check_field(category, field_name):
    """Yeni kontrol alanı ekler"""
    wb = get_excel_file()
    if category not in wb.sheetnames:
        wb.create_sheet(category)
        ws = wb[category]
        ws.append(["Field"])
    else:
        ws = wb[category]
    # Aynı alan var mı kontrol et
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == field_name:
            return False
    ws.append([field_name])
    wb.save(EXCEL_FILE)
    return True

def delete_check_field(category, field_name):
    """Kontrol alanını siler"""
    wb = get_excel_file()
    if category not in wb.sheetnames:
        return False
    ws = wb[category]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == field_name:
            ws.delete_rows(row_idx)
            wb.save(EXCEL_FILE)
            return True
    return False

def update_check_field(category, old_name, new_name):
    """Kontrol alanını günceller"""
    wb = get_excel_file()
    if category not in wb.sheetnames:
        return False
    ws = wb[category]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == old_name:
            ws.cell(row=row_idx, column=1, value=new_name)
            wb.save(EXCEL_FILE)
            return True
    return False

def add_item(item_name):
    """Yeni eşya ekler"""
    wb = get_excel_file()
    ws = wb["Items"]
    # Aynı eşya var mı kontrol et
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == item_name:
            return False
    ws.append([item_name])
    wb.save(EXCEL_FILE)
    return True

def delete_item(item_name):
    """Eşyayı siler"""
    wb = get_excel_file()
    ws = wb["Items"]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == item_name:
            ws.delete_rows(row_idx)
            wb.save(EXCEL_FILE)
            return True
    return False

def update_item(old_name, new_name):
    """Eşya adını günceller"""
    wb = get_excel_file()
    ws = wb["Items"]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == old_name:
            ws.cell(row=row_idx, column=1, value=new_name)
            wb.save(EXCEL_FILE)
            return True
    return False

def add_user(username, password, full_name, email="", is_admin_user=False):
    """Yeni kullanıcı ekler"""
    # Google Sheets kullanılıyorsa
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Users")
                all_values = sheet.get_all_values()
                
                # Başlık kontrolü
                if not all_values:
                    sheet.append_row(["Username", "Password", "Full Name", "Email", "Admin"])
                
                # Kullanıcı zaten var mı kontrol et
                for row in all_values[1:]:
                    if row and row[0] == username:
                        return False
                
                # Yeni kullanıcı ekle
                sheet.append_row([username, password, full_name, email, "Yes" if is_admin_user else "No"])
                return True
            except Exception as e:
                _log("E", "excel_handler.py:add_user:google_sheets", "Failed to add user to Google Sheets", {"error": str(e)})
                return False
    
    # Excel'e ekle
    wb = get_excel_file()
    ws = wb["Users"]
    
    # Başlık satırını kontrol et ve gerekli kolonları ekle
    headers = [cell.value for cell in ws[1]]
    if "Email" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Email")
        headers.append("Email")
    if "Admin" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Admin")
        headers.append("Admin")
    
    # Kullanıcı zaten var mı kontrol et
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == username:
            return False
    
    # Yeni satır ekle
    username_col = 1
    password_col = 2
    full_name_col = 3
    email_col = headers.index("Email") + 1 if "Email" in headers else 4
    admin_col = headers.index("Admin") + 1 if "Admin" in headers else 5
    
    new_row = [None] * max(username_col, password_col, full_name_col, email_col, admin_col)
    new_row[username_col - 1] = username
    new_row[password_col - 1] = password
    new_row[full_name_col - 1] = full_name
    if email_col <= len(new_row):
        new_row[email_col - 1] = email
    if admin_col <= len(new_row):
        new_row[admin_col - 1] = "Yes" if is_admin_user else "No"
    
    ws.append(new_row)
    wb.save(EXCEL_FILE)
    return True

def delete_user(username):
    """Kullanıcıyı siler"""
    # Google Sheets kullanılıyorsa
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Users")
                all_values = sheet.get_all_values()
                
                if not all_values or len(all_values) < 2:
                    return False
                
                # Kullanıcıyı bul ve sil
                for i, row in enumerate(all_values[1:], start=2):
                    if row and row[0] == username:
                        sheet.delete_rows(i)
                        return True
                
                return False
            except Exception as e:
                _log("E", "excel_handler.py:delete_user:google_sheets", "Failed to delete user from Google Sheets", {"error": str(e)})
                return False
    
    # Excel'den sil
    try:
        wb = get_excel_file()
        ws = wb["Users"]
        
        # Kullanıcıyı bul ve sil
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == username:
                ws.delete_rows(row_idx)
                wb.save(EXCEL_FILE)
                return True
        
        return False
    except Exception as e:
        _log("E", "excel_handler.py:delete_user", "Failed to delete user", {"error": str(e)})
        return False

def update_user(username, password=None, full_name=None, email=None, is_admin=None):
    """Kullanıcı bilgilerini günceller"""
    # Google Sheets kullanılıyorsa
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Users")
                all_values = sheet.get_all_values()
                
                if not all_values or len(all_values) < 2:
                    return False
                
                headers = all_values[0]
                
                # Kullanıcıyı bul ve güncelle
                for i, row in enumerate(all_values[1:], start=2):
                    if row and row[0] == username:
                        if password is not None:
                            pwd_col = headers.index("Password") + 1 if "Password" in headers else 2
                            sheet.update_cell(i, pwd_col, password)
                        if full_name is not None:
                            name_col = headers.index("Full Name") + 1 if "Full Name" in headers else 3
                            sheet.update_cell(i, name_col, full_name)
                        if email is not None:
                            email_col = headers.index("Email") + 1 if "Email" in headers else 4
                            sheet.update_cell(i, email_col, email)
                        if is_admin is not None:
                            admin_col = headers.index("Admin") + 1 if "Admin" in headers else 5
                            sheet.update_cell(i, admin_col, "Yes" if is_admin else "No")
                        return True
                
                return False
            except Exception as e:
                _log("E", "excel_handler.py:update_user:google_sheets", "Failed to update user in Google Sheets", {"error": str(e)})
                return False
    
    # Excel'den güncelle
    try:
        wb = get_excel_file()
        ws = wb["Users"]
        
        headers = [cell.value for cell in ws[1]]
        
        # Kullanıcıyı bul ve güncelle
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == username:
                if password is not None:
                    pwd_col = headers.index("Password") + 1 if "Password" in headers else 2
                    ws.cell(row=row_idx, column=pwd_col, value=password)
                if full_name is not None:
                    name_col = headers.index("Full Name") + 1 if "Full Name" in headers else 3
                    ws.cell(row=row_idx, column=name_col, value=full_name)
                if email is not None:
                    email_col = headers.index("Email") + 1 if "Email" in headers else 4
                    ws.cell(row=row_idx, column=email_col, value=email)
                if is_admin is not None:
                    admin_col = headers.index("Admin") + 1 if "Admin" in headers else 5
                    ws.cell(row=row_idx, column=admin_col, value="Yes" if is_admin else "No")
                
                wb.save(EXCEL_FILE)
                return True
        
        return False
    except Exception as e:
        _log("E", "excel_handler.py:update_user", "Failed to update user", {"error": str(e)})
        return False

def update_excel_with_admin_column():
    """Mevcut Excel dosyasına Admin ve Email kolonlarını ekler ve admin kullanıcısını ekler
    Google Sheets kullanılıyorsa bu fonksiyon hiçbir şey yapmaz (Google Sheets'te manuel yapılmalı)
    """
    # Google Sheets kullanılıyorsa Excel işlemlerini atla
    if USE_GOOGLE_SHEETS:
        _log("D", "excel_handler.py:update_excel_with_admin_column", "Google Sheets enabled, skipping Excel update", {})
        return
    
    try:
        wb = get_excel_file()
        ws = wb["Users"]
        
        # Başlık satırını kontrol et
        headers = [cell.value for cell in ws[1]]
        
        # Email kolonu yoksa ekle
        if "Email" not in headers:
            _log("D", "excel_handler.py:update_excel_with_admin_column", "Adding Email column to headers", {"current_headers": headers})
            # Email kolonunu Full Name'den sonra ekle
            email_col_idx = 4 if len(headers) >= 3 else len(headers) + 1
            ws.insert_cols(email_col_idx)
            ws.cell(row=1, column=email_col_idx, value="Email")
            headers.insert(email_col_idx - 1, "Email")
            
            # Mevcut kullanıcılara boş email ekle
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=email_col_idx, value="")
        
        # Admin kolonu yoksa ekle
        if "Admin" not in headers:
            _log("D", "excel_handler.py:update_excel_with_admin_column", "Adding Admin column to headers", {"current_headers": headers})
            ws.cell(row=1, column=len(headers) + 1, value="Admin")
            headers.append("Admin")
            
            # Mevcut kullanıcılara "No" ekle
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=len(headers), value="No")
        
        # Admin kullanıcısı var mı kontrol et
        admin_exists = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == "admin":
                admin_exists = True
                break
        
        # Admin kullanıcısı yoksa ekle
        if not admin_exists:
            _log("D", "excel_handler.py:update_excel_with_admin_column", "Adding admin user", {})
            ws.append(["admin", "admin123", "Admin User", "admin@example.com", "Yes"])
        
        wb.save(EXCEL_FILE)
        _log("D", "excel_handler.py:update_excel_with_admin_column", "Excel updated successfully", {})
    except Exception as e:
        # Excel dosyası bozuksa veya oluşturulamazsa hata verme, sadece log
        _log("E", "excel_handler.py:update_excel_with_admin_column", "Failed to update Excel file", {"error": str(e)})
        # Bulut ortamında Excel dosyası olmayabilir, bu normal
        pass

# Form gönderimleri için form_data.xlsx dosyasındaki Submissions sheet'i kullanılacak
# Önce mevcut dizinde ara, yoksa geçici dizin kullan
if os.path.exists(EXCEL_FILE_LOCAL):
    SUBMISSIONS_FILE = EXCEL_FILE_LOCAL
else:
    SUBMISSIONS_FILE = EXCEL_FILE_TEMP

def _prepare_submission_row(form_data):
    """Form verilerini Excel/Sheets satırına dönüştürür"""
    from datetime import datetime
    
    # Başlık satırı oluştur
    headers = [
        "Timestamp", "Driver Name", "Vehicle", "Odometer Start", 
        "Fuel Level", "Oil Level", "Fuel Card", "Measuring Tape", 
        "Safety Vest", "Fuel Amount", "Additional Comments"
    ]
    
    # Exterior checks başlıkları
    exterior_fields = load_check_fields("ExteriorChecks")
    for field in exterior_fields:
        headers.append(f"Exterior_{field}")
    
    # Engine checks başlıkları
    engine_fields = load_check_fields("EngineChecks")
    for field in engine_fields:
        headers.append(f"Engine_{field}")
    
    # Safety equipment başlıkları
    safety_fields = load_check_fields("SafetyEquipment")
    for field in safety_fields:
        headers.append(f"Safety_{field}")
    
    # Interior checks başlıkları
    interior_fields = load_check_fields("InteriorChecks")
    for field in interior_fields:
        headers.append(f"Interior_{field}")
    
    # Veri satırı oluştur
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        form_data.get("driver_name", ""),
        form_data.get("vehicle", ""),
        form_data.get("odometer_start", ""),
        form_data.get("fuel_level", ""),
        form_data.get("oil_level", ""),
        form_data.get("fuel_card", ""),
        form_data.get("measuring_tape", ""),
        form_data.get("safety_vest", ""),
        form_data.get("fuel_amount", ""),
        form_data.get("additional_comments", "")
    ]
    
    # Exterior checks değerleri
    exterior_checks = form_data.get("exterior_checks", {})
    for field in exterior_fields:
        row.append(exterior_checks.get(field, ""))
    
    # Engine checks değerleri
    engine_checks = form_data.get("engine_checks", {})
    for field in engine_fields:
        row.append(engine_checks.get(field, ""))
    
    # Safety equipment değerleri
    safety_checks = form_data.get("safety_checks", {})
    for field in safety_fields:
        row.append(safety_checks.get(field, ""))
    
    # Interior checks değerleri
    interior_checks = form_data.get("interior_checks", {})
    for field in interior_fields:
        row.append(interior_checks.get(field, ""))
    
    return headers, row

def save_form_submission_to_google_apps_script(form_data):
    """Form verilerini Google Apps Script'e HTTP POST ile gönderir (eski yöntem - Google Sheets formatına uygun)"""
    import urllib.request
    import urllib.parse
    import json
    
    try:
        # Form verilerini Google Sheets formatına uygun şekilde düzleştir
        # Eski HTML formundaki field isimlerini kullan
        flat_data = {}
        
        # Temel alanlar (eski formdaki isimlerle aynı)
        flat_data["driver_name"] = form_data.get("driver_name", "")
        
        # Vehicle - "Other" seçildiyse other_vehicle kullan
        vehicle = form_data.get("vehicle", "")
        if vehicle == "Other":
            flat_data["vehicle"] = "Other"
            flat_data["other_vehicle"] = form_data.get("other_vehicle", "")
        else:
            flat_data["vehicle"] = vehicle
            flat_data["other_vehicle"] = ""
        
        flat_data["odometer_start"] = str(form_data.get("odometer_start", ""))
        
        # Fuel level - "Other" seçildiyse other_fuel kullan
        fuel_level = form_data.get("fuel_level", "")
        if fuel_level == "Other":
            flat_data["fuel_level"] = "Other"
            flat_data["other_fuel"] = form_data.get("other_fuel", "")
        else:
            flat_data["fuel_level"] = fuel_level
            flat_data["other_fuel"] = ""
        
        flat_data["oil_level"] = form_data.get("oil_level", "")
        flat_data["fuel_card"] = form_data.get("fuel_card", "")
        flat_data["measuring_tape"] = form_data.get("measuring_tape", "")
        flat_data["safety_vest"] = form_data.get("safety_vest", "")
        flat_data["fuel_amount"] = form_data.get("fuel_amount", "")
        flat_data["additional_comments"] = form_data.get("additional_comments", "")
        
        # Dosya yüklemeleri (şimdilik boş, gelecekte eklenebilir)
        flat_data["odometer_file"] = ""
        flat_data["fuel_receipt"] = ""
        
        # Exterior checks - direkt field isimleri (prefix olmadan)
        exterior_checks = form_data.get("exterior_checks", {})
        exterior_fields = ['headlights', 'break_lights', 'indicators', 'mirrors', 'windows', 
                          'windshield', 'wiper_fluid', 'wipers', 'tires', 'body_paint']
        for field in exterior_fields:
            flat_data[field] = exterior_checks.get(field, "Needs Attention")
        
        # Engine checks - direkt field isimleri
        engine_checks = form_data.get("engine_checks", {})
        engine_fields = ['engine_noise', 'coolant_level', 'brake_fluid', 'battery_condition', 'belts_hoses']
        for field in engine_fields:
            flat_data[field] = engine_checks.get(field, "Needs Attention")
        
        # Safety equipment - direkt field isimleri
        safety_checks = form_data.get("safety_checks", {})
        safety_fields = ['seatbelts', 'fire_extinguisher', 'first_aid_kit', 'horn', 'jumper_cables', 'snow_brush']
        for field in safety_fields:
            flat_data[field] = safety_checks.get(field, "Needs Attention")
        
        # Interior checks - direkt field isimleri
        interior_checks = form_data.get("interior_checks", {})
        interior_fields = ['cleanliness', 'dashboard_lights', 'hvac', 'seats']
        for field in interior_fields:
            flat_data[field] = interior_checks.get(field, "Needs Attention")
        
        # HTTP POST isteği gönder (multipart/form-data yerine application/x-www-form-urlencoded)
        data = urllib.parse.urlencode(flat_data).encode('utf-8')
        req = urllib.request.Request(GOOGLE_APPS_SCRIPT_URL, data=data, method='POST')
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')
        
        with urllib.request.urlopen(req, timeout=10) as response:
            result = response.read().decode('utf-8')
            # #region agent log
            _log("F", "excel_handler.py:save_form_submission_to_google_apps_script", "Google Apps Script response", {"status_code": response.status, "result": result[:100]})
            # #endregion agent log
            return True
    except Exception as e:
        # #region agent log
        _log("F", "excel_handler.py:save_form_submission_to_google_apps_script", "Google Apps Script error", {"error": str(e)})
        # #endregion agent log
        return False

def save_form_submission(form_data):
    """Form verilerini Excel dosyasına veya Google Sheets'e kaydeder"""
    from datetime import datetime
    from openpyxl import Workbook
    
    headers, row = _prepare_submission_row(form_data)
    
    # Google Apps Script kullanılıyorsa (eski yöntem - öncelikli)
    if USE_GOOGLE_APPS_SCRIPT and GOOGLE_APPS_SCRIPT_URL:
        success = save_form_submission_to_google_apps_script(form_data)
        if success:
            # #region agent log
            _log("F", "excel_handler.py:save_form_submission", "Saved to Google Apps Script successfully", {})
            # #endregion agent log
            # Yerel Excel'e de kaydet (backup)
            pass  # Fall through to Excel save
    
    # Google Sheets kullanılıyorsa (yeni yöntem)
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Submissions")
                # Başlık satırını kontrol et
                existing_headers = sheet.row_values(1)
                if not existing_headers or len(existing_headers) < len(headers):
                    sheet.clear()
                    sheet.append_row(headers)
                # Yeni satır ekle
                sheet.append_row(row)
                return
            except Exception as e:
                # #region agent log
                _log("E", "excel_handler.py:save_form_submission:google_sheets", "Google Sheets save failed, falling back to Excel", {"error": str(e)})
                # #endregion agent log
                pass  # Fallback to Excel
    
    # Excel dosyasına kaydet (form_data.xlsx içindeki Submissions sheet'ine)
    wb = get_excel_file()
    
    # Submissions sheet'i yoksa oluştur
    if "Submissions" not in wb.sheetnames:
        ws = wb.create_sheet("Submissions")
        ws.append(headers)
    else:
        ws = wb["Submissions"]
        # Başlık satırı yoksa ekle
        if ws.max_row == 0 or not any(ws.cell(row=1, column=col).value for col in range(1, len(headers) + 1)):
            ws.append(headers)
    
    ws.append(row)
    wb.save(EXCEL_FILE)

def load_form_submissions():
    """Form gönderimlerini Excel'den veya Google Sheets'ten okur"""
    # Google Sheets kullanılıyorsa
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Submissions")
                all_values = sheet.get_all_values()
                
                if not all_values or len(all_values) < 2:
                    return []
                
                headers = all_values[0]
                submissions = []
                
                for row in all_values[1:]:
                    if row and row[0]:  # Timestamp varsa
                        submission = {}
                        for i, header in enumerate(headers):
                            submission[header] = row[i] if i < len(row) else None
                        submissions.append(submission)
                
                return submissions
            except Exception as e:
                # #region agent log
                _log("E", "excel_handler.py:load_form_submissions:google_sheets", "Google Sheets load failed, falling back to Excel", {"error": str(e)})
                # #endregion agent log
                pass  # Fallback to Excel
    
    # Excel dosyasından oku (form_data.xlsx içindeki Submissions sheet'inden)
    wb = get_excel_file()
    if "Submissions" not in wb.sheetnames:
        return []
    
    ws = wb["Submissions"]
    submissions = []
    
    # Başlık satırını oku
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # Veri satırlarını oku
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Timestamp varsa
            submission = {}
            for i, header in enumerate(headers):
                submission[header] = row[i] if i < len(row) else None
            submissions.append(submission)
    
    return submissions

def is_admin(username):
    """Kullanıcının admin olup olmadığını kontrol eder"""
    _log("B", "excel_handler.py:is_admin:entry", "is_admin called", {"username": username})
    
    # Google Sheets'ten oku
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Users")
                all_values = sheet.get_all_values()
                
                if not all_values or len(all_values) < 2:
                    return False
                
                headers = all_values[0]
                _log("B", "excel_handler.py:is_admin:headers", "Users sheet headers", {"headers": headers, "has_admin": "Admin" in headers})
                
                admin_col_idx = None
                if "Admin" in headers:
                    admin_col_idx = headers.index("Admin")
                elif len(headers) > 3:
                    admin_col_idx = 3
                
                for row in all_values[1:]:
                    if row and row[0] == username:
                        if admin_col_idx and len(row) > admin_col_idx:
                            admin_value = row[admin_col_idx]
                            is_admin_result = admin_value == "Yes" or admin_value == True
                            _log("B", "excel_handler.py:is_admin:admin_check", "Admin value check", {"admin_value": admin_value, "is_admin": is_admin_result})
                            return is_admin_result
                        if username == "admin":
                            return True
                
                _log("B", "excel_handler.py:is_admin:user_not_found", "User not found or not admin", {"username": username})
                return False
            except Exception as e:
                _log("E", "excel_handler.py:is_admin:google_sheets", "Google Sheets load failed, falling back to Excel", {"error": str(e)})
    
    # Excel'den oku (fallback)
    wb = get_excel_file()
    ws = wb["Users"]
    
    headers = [cell.value for cell in ws[1]]
    _log("B", "excel_handler.py:is_admin:headers", "Users sheet headers", {"headers": headers, "has_admin": "Admin" in headers})
    admin_col_idx = None
    
    if "Admin" in headers:
        admin_col_idx = headers.index("Admin")
        _log("B", "excel_handler.py:is_admin:admin_col_found", "Admin column found", {"admin_col_idx": admin_col_idx})
    elif len(headers) > 3:
        admin_col_idx = 3
        _log("B", "excel_handler.py:is_admin:admin_col_assumed", "Assuming admin column at index 3", {"admin_col_idx": admin_col_idx})
    
    user_found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        _log("B", "excel_handler.py:is_admin:checking_row", "Checking row", {"row_username": row[0] if row else None, "matches": row[0] == username if row else False})
        if row[0] == username:
            user_found = True
            _log("B", "excel_handler.py:is_admin:user_found", "User found in sheet", {"username": username, "row_length": len(row), "admin_col_idx": admin_col_idx})
            if admin_col_idx and len(row) > admin_col_idx:
                admin_value = row[admin_col_idx]
                is_admin_result = admin_value == "Yes" or admin_value == True
                _log("B", "excel_handler.py:is_admin:admin_check", "Admin value check", {"admin_value": admin_value, "is_admin": is_admin_result})
                return is_admin_result
            if username == "admin":
                _log("B", "excel_handler.py:is_admin:fallback", "Using fallback admin check", {"username": username, "is_admin": True})
                return True
    
    _log("B", "excel_handler.py:is_admin:user_not_found", "User not found or not admin", {"username": username, "user_found": user_found})
    return False

# Şifre sıfırlama kodları için geçici dosya
RESET_CODES_FILE = os.path.join(TEMP_DIR, "reset_codes.json")

def get_user_by_email(email):
    """E-posta adresine göre kullanıcı bilgilerini döndürür"""
    users = load_users()
    for username, user_data in users.items():
        if user_data.get("email", "").lower() == email.lower():
            return username, user_data
    return None, None

def verify_user_email(username, email):
    """Kullanıcı adı ve e-posta kombinasyonunu doğrular"""
    users = load_users()
    user_data = users.get(username)
    if user_data and user_data.get("email", "").lower() == email.lower():
        return True, user_data
    return False, None

def generate_reset_code():
    """6 haneli rastgele şifre sıfırlama kodu oluşturur"""
    import random
    return str(random.randint(100000, 999999))

def save_reset_code(email, code, username):
    """Şifre sıfırlama kodunu kaydeder (10 dakika geçerli)"""
    import json
    from datetime import datetime, timedelta
    
    codes = {}
    if os.path.exists(RESET_CODES_FILE):
        try:
            with open(RESET_CODES_FILE, 'r', encoding='utf-8') as f:
                codes = json.load(f)
        except:
            codes = {}
    
    # Eski kodları temizle
    now = datetime.now()
    codes = {k: v for k, v in codes.items() if datetime.fromisoformat(v['expires']) > now}
    
    # Yeni kodu ekle
    expires = (now + timedelta(minutes=10)).isoformat()
    codes[code] = {
        'email': email.lower(),
        'username': username,
        'expires': expires
    }
    
    with open(RESET_CODES_FILE, 'w', encoding='utf-8') as f:
        json.dump(codes, f, ensure_ascii=False, indent=2)
    
    return True

def verify_reset_code(code):
    """Şifre sıfırlama kodunu doğrular ve kullanıcı bilgisini döndürür"""
    import json
    from datetime import datetime
    
    if not os.path.exists(RESET_CODES_FILE):
        return None, None
    
    try:
        with open(RESET_CODES_FILE, 'r', encoding='utf-8') as f:
            codes = json.load(f)
    except:
        return None, None
    
    if code not in codes:
        return None, None
    
    code_data = codes[code]
    expires = datetime.fromisoformat(code_data['expires'])
    
    if datetime.now() > expires:
        # Süresi dolmuş, kodu sil
        del codes[code]
        with open(RESET_CODES_FILE, 'w', encoding='utf-8') as f:
            json.dump(codes, f, ensure_ascii=False, indent=2)
        return None, None
    
    return code_data['username'], code_data['email']

def send_reset_code_email(email, code):
    """Şifre sıfırlama kodunu e-posta ile gönderir"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    # E-posta ayarları (Streamlit secrets'tan veya environment variable'dan)
    smtp_server = get_secret("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(get_secret("SMTP_PORT", "587"))
    smtp_username = get_secret("SMTP_USERNAME", "")
    smtp_password = get_secret("SMTP_PASSWORD", "")
    
    # E-posta ayarları yoksa, konsola yazdır (geliştirme için)
    if not smtp_username or not smtp_password:
        # Windows konsolu encoding sorunu için ASCII-safe mesaj
        print(f"[DEBUG] Reset code for {email}: {code}")
        return True  # Geliştirme modunda başarılı say
    
    try:
        # E-posta oluştur
        msg = MIMEMultipart()
        msg['From'] = smtp_username
        msg['To'] = email
        msg['Subject'] = "Şifre Sıfırlama Kodu - Araç Kontrol Formu"
        
        body = f"""
Merhaba,

Araç Kontrol Formu uygulaması için şifre sıfırlama talebiniz alınmıştır.

Şifre sıfırlama kodunuz: {code}

Bu kod 10 dakika süreyle geçerlidir.

Eğer bu talebi siz yapmadıysanız, lütfen bu e-postayı görmezden gelin.

İyi günler,
Araç Kontrol Formu Ekibi
        """
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # E-posta gönder
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)
        server.quit()
        
        return True
    except Exception as e:
        _log("E", "excel_handler.py:send_reset_code_email", "Failed to send email", {"error": str(e)})
        # Hata olsa bile geliştirme modunda devam et
        print(f"[DEBUG] Email could not be sent, code: {code}")
        return True

def update_user_password(username, new_password):
    """Kullanıcı şifresini günceller"""
    # Google Sheets kullanılıyorsa
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Users")
                all_values = sheet.get_all_values()
                
                if not all_values or len(all_values) < 2:
                    return False
                
                headers = all_values[0]
                password_col_idx = headers.index("Password") if "Password" in headers else 1
                
                # Kullanıcıyı bul ve şifresini güncelle
                for i, row in enumerate(all_values[1:], start=2):
                    if row and row[0] == username:
                        sheet.update_cell(i, password_col_idx + 1, new_password)
                        return True
                
                return False
            except Exception as e:
                _log("E", "excel_handler.py:update_user_password:google_sheets", "Failed to update password in Google Sheets", {"error": str(e)})
                return False
    
    # Excel'den güncelle
    try:
        wb = get_excel_file()
        ws = wb["Users"]
        
        headers = [cell.value for cell in ws[1]]
        password_col_idx = headers.index("Password") if "Password" in headers else 1
        
        # Kullanıcıyı bul ve şifresini güncelle
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == username:
                ws.cell(row=row_idx, column=password_col_idx + 1, value=new_password)
                wb.save(EXCEL_FILE)
                return True
        
        return False
    except Exception as e:
        _log("E", "excel_handler.py:update_user_password", "Failed to update password", {"error": str(e)})
        return False

def delete_reset_code(code):
    """Kullanılan şifre sıfırlama kodunu siler"""
    import json
    
    if not os.path.exists(RESET_CODES_FILE):
        return
    
    try:
        with open(RESET_CODES_FILE, 'r', encoding='utf-8') as f:
            codes = json.load(f)
        
        if code in codes:
            del codes[code]
            with open(RESET_CODES_FILE, 'w', encoding='utf-8') as f:
                json.dump(codes, f, ensure_ascii=False, indent=2)
    except:
        pass

def update_user_email(username, email):
    """Kullanıcının e-posta adresini günceller"""
    # Google Sheets kullanılıyorsa
    if USE_GOOGLE_SHEETS:
        client = get_google_sheets_client()
        if client:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Users")
                all_values = sheet.get_all_values()
                
                if not all_values or len(all_values) < 2:
                    return False
                
                headers = all_values[0]
                email_col_idx = headers.index("Email") if "Email" in headers else 3
                
                # Kullanıcıyı bul ve e-postasını güncelle
                for i, row in enumerate(all_values[1:], start=2):
                    if row and row[0] == username:
                        sheet.update_cell(i, email_col_idx + 1, email)
                        return True
                
                return False
            except Exception as e:
                _log("E", "excel_handler.py:update_user_email:google_sheets", "Failed to update email in Google Sheets", {"error": str(e)})
                return False
    
    # Excel'den güncelle
    try:
        wb = get_excel_file()
        ws = wb["Users"]
        
        headers = [cell.value for cell in ws[1]]
        email_col_idx = headers.index("Email") if "Email" in headers else 3
        
        # Kullanıcıyı bul ve e-postasını güncelle
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == username:
                ws.cell(row=row_idx, column=email_col_idx + 1, value=email)
                wb.save(EXCEL_FILE)
                return True
        
        return False
    except Exception as e:
        _log("E", "excel_handler.py:update_user_email", "Failed to update email", {"error": str(e)})
        return False

